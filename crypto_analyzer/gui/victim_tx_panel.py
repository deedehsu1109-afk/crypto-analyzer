from __future__ import annotations
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk

from database import db as _db
from api.price_fetcher import fetch_exchange_rate, CURRENCY_MAP

_COLS = [
    ("日期",       "tx_date",       100),
    ("時間(UTC+8)", "tx_time",       90),
    ("FROM（發起錢包）", "from_addr", 220),
    ("TO（接收錢包）",  "to_addr",    220),
    ("金額(NT)",   "amount_ntd",    100),
    ("數量",       "quantity",      80),
    ("幣種",       "currency",      65),
    ("交易匯率",   "exchange_rate", 90),
    ("當日均價",   "daily_avg",     90),
]

_CURRENCIES = sorted(CURRENCY_MAP.keys()) + ["OTHER"]


# ── 單筆新增 / 編輯對話框 ──────────────────────────────────────────────────────

class VictimTxDialog(ctk.CTkToplevel):
    def __init__(self, parent, case_id: int, row: dict = None, on_save=None):
        super().__init__(parent)
        self.case_id = case_id
        self.row     = row or {}
        self.on_save = on_save
        self.title("編輯交易記錄" if row else "新增交易記錄")
        self.geometry("620x560")
        self.resizable(False, False)
        self.grab_set()
        self._build()
        if row:
            self._fill(row)

    def _lbl(self, parent, text, row, required=False):
        color = "#ff9999" if required else None
        ctk.CTkLabel(parent, text=text,
                     font=("Microsoft JhengHei", 11, "bold"),
                     text_color=color,
                     anchor="e", width=120).grid(
            row=row, column=0, padx=(12, 4), pady=5, sticky="e")

    def _entry(self, parent, row, placeholder="", width=220) -> ctk.CTkEntry:
        e = ctk.CTkEntry(parent, font=("Consolas", 11),
                         placeholder_text=placeholder, width=width)
        e.grid(row=row, column=1, padx=(4, 12), pady=5, sticky="w")
        return e

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        f = ctk.CTkFrame(self, corner_radius=10)
        f.grid(row=0, column=0, sticky="nsew", padx=14, pady=14)
        f.grid_columnconfigure(1, weight=1)

        # 日期
        self._lbl(f, "日期*：", 0, True)
        self.date_e = self._entry(f, 0, "YYYY-MM-DD")

        # 時間
        self._lbl(f, "時間 (UTC+8)*：", 1, True)
        self.time_e = self._entry(f, 1, "HH:MM:SS")

        # FROM
        self._lbl(f, "FROM（發起）：", 2)
        self.from_e = self._entry(f, 2, "發起錢包地址（全碼）", width=380)
        self.from_e.configure(width=380)

        # TO
        self._lbl(f, "TO（接收）：", 3)
        self.to_e = self._entry(f, 3, "接收錢包地址（全碼）", width=380)
        self.to_e.configure(width=380)

        # 金額 NT
        self._lbl(f, "金額(NT)*：", 4, True)
        amt_frame = ctk.CTkFrame(f, fg_color="transparent")
        amt_frame.grid(row=4, column=1, sticky="w", pady=5)
        self.amt_e = ctk.CTkEntry(amt_frame, font=("Consolas", 11),
                                   placeholder_text="被害人陳述之新台幣金額",
                                   width=160)
        self.amt_e.pack(side="left")
        ctk.CTkLabel(amt_frame, text="元",
                     font=("Microsoft JhengHei", 10),
                     text_color="gray70").pack(side="left", padx=(4, 0))

        # 幣種
        self._lbl(f, "幣種*：", 5, True)
        cur_frame = ctk.CTkFrame(f, fg_color="transparent")
        cur_frame.grid(row=5, column=1, sticky="w", pady=5)
        self.cur_var = ctk.StringVar(value="USDT")
        ctk.CTkOptionMenu(cur_frame, variable=self.cur_var,
                          values=_CURRENCIES, width=100,
                          font=("Consolas", 11)).pack(side="left")
        ctk.CTkLabel(cur_frame, text="或輸入：",
                     font=("Microsoft JhengHei", 10)).pack(side="left", padx=(8, 2))
        self.cur_other = ctk.CTkEntry(cur_frame, width=80,
                                      font=("Consolas", 11),
                                      placeholder_text="自訂幣種")
        self.cur_other.pack(side="left")

        # 數量
        self._lbl(f, "數量*：", 6, True)
        qty_frame = ctk.CTkFrame(f, fg_color="transparent")
        qty_frame.grid(row=6, column=1, sticky="w", pady=5)
        self.qty_e = ctk.CTkEntry(qty_frame, width=160, font=("Consolas", 11),
                                   placeholder_text="被害人陳述之幣種數量")
        self.qty_e.pack(side="left")
        ctk.CTkLabel(qty_frame, text="（幣）",
                     font=("Microsoft JhengHei", 10),
                     text_color="gray70").pack(side="left", padx=(4, 0))

        # 交易匯率（由金額/數量自動計算，可手動覆寫）
        ctk.CTkLabel(f,
                     text="交易匯率\n(NT/幣)：",
                     font=("Microsoft JhengHei", 11, "bold"),
                     text_color="#ffcc55",
                     anchor="e", width=120).grid(
            row=7, column=0, padx=(12, 4), pady=5, sticky="e")
        rate_frame = ctk.CTkFrame(f, fg_color="transparent")
        rate_frame.grid(row=7, column=1, sticky="w", pady=5)
        self.rate_e = ctk.CTkEntry(rate_frame, font=("Consolas", 11),
                                    width=140,
                                    placeholder_text="自動計算")
        self.rate_e.pack(side="left")
        self._rate_hint = ctk.CTkLabel(
            rate_frame,
            text="= 金額(NT) ÷ 數量",
            font=("Microsoft JhengHei", 9),
            text_color="gray60")
        self._rate_hint.pack(side="left", padx=(8, 0))

        # 當日均價（市場行情參考，由 CoinGecko API 查詢）
        ctk.CTkLabel(f,
                     text="當日均價\n(市場行情)：",
                     font=("Microsoft JhengHei", 11, "bold"),
                     text_color="#aaffaa",
                     anchor="e", width=120).grid(
            row=8, column=0, padx=(12, 4), pady=5, sticky="e")
        avg_frame = ctk.CTkFrame(f, fg_color="transparent")
        avg_frame.grid(row=8, column=1, sticky="w", pady=5)
        self.avg_e = ctk.CTkEntry(avg_frame, font=("Consolas", 11),
                                   width=140,
                                   placeholder_text="自動查詢")
        self.avg_e.pack(side="left")
        self._fetch_btn = ctk.CTkButton(avg_frame,
                                        text="查詢市場行情",
                                        width=110,
                                        font=("Microsoft JhengHei", 10),
                                        fg_color="#2a4a8a",
                                        command=self._fetch_price)
        self._fetch_btn.pack(side="left", padx=(10, 0))
        ctk.CTkLabel(avg_frame,
                     text="（最高+最低）÷2，來源 CoinGecko",
                     font=("Microsoft JhengHei", 9),
                     text_color="gray60").pack(side="left", padx=(6, 0))

        # 當日高低（唯讀）
        self._lbl(f, "當日高/低(NT)：", 9)
        self.hl_lbl = ctk.CTkLabel(f, text="—",
                                   font=("Consolas", 10), text_color="gray60")
        self.hl_lbl.grid(row=9, column=1, padx=(4, 12), pady=2, sticky="w")

        # 綁定自動計算交易匯率
        self.amt_e.bind("<KeyRelease>", self._auto_calc_rate)
        self.qty_e.bind("<KeyRelease>", self._auto_calc_rate)

        # 備註
        self._lbl(f, "備註：", 10)
        self.notes_e = ctk.CTkEntry(f, width=380, font=("Microsoft JhengHei", 11))
        self.notes_e.grid(row=10, column=1, padx=(4, 12), pady=5, sticky="w")

        # 按鈕
        btn_f = ctk.CTkFrame(self, fg_color="transparent")
        btn_f.grid(row=1, column=0, pady=(0, 14))
        ctk.CTkButton(btn_f, text="儲存", width=110,
                      font=("Microsoft JhengHei", 12, "bold"),
                      command=self._save).pack(side="left", padx=8)
        ctk.CTkButton(btn_f, text="取消", width=80,
                      fg_color="gray40", font=("Microsoft JhengHei", 11),
                      command=self.destroy).pack(side="left", padx=4)

    def _fill(self, row: dict):
        def _set(entry, key):
            v = str(row.get(key) or "")
            if v and v != "None":
                entry.delete(0, "end")
                entry.insert(0, v)
        _set(self.date_e, "tx_date")
        _set(self.time_e, "tx_time")
        _set(self.from_e, "from_addr")
        _set(self.to_e,   "to_addr")
        _set(self.amt_e,  "amount_ntd")
        _set(self.qty_e,  "quantity")
        _set(self.rate_e, "exchange_rate")
        _set(self.avg_e,  "daily_avg")
        _set(self.notes_e,"notes")
        cur = str(row.get("currency") or "USDT")
        if cur in _CURRENCIES:
            self.cur_var.set(cur)
        else:
            self.cur_var.set("OTHER")
            self.cur_other.insert(0, cur)
        h = row.get("daily_high")
        l = row.get("daily_low")
        if h and l:
            self.hl_lbl.configure(text=f"高 {h:,.2f} ／ 低 {l:,.2f}")

    def _get_currency(self) -> str:
        c = self.cur_other.get().strip().upper()
        return c if c else self.cur_var.get()

    def _auto_calc_rate(self, _event=None):
        """金額或數量變動時，自動計算交易匯率 = 金額(NT) / 數量"""
        amt_s = self.amt_e.get().strip().replace(",", "")
        qty_s = self.qty_e.get().strip().replace(",", "")
        try:
            amt = float(amt_s) if amt_s else None
            qty = float(qty_s) if qty_s else None
            if amt and qty and qty > 0:
                rate = round(amt / qty, 4)
                self.rate_e.delete(0, "end")
                self.rate_e.insert(0, str(rate))
                self._rate_hint.configure(
                    text=f"= {amt:,.0f} ÷ {qty:,.6f} = {rate:,.4f}",
                    text_color="#ffcc55")
            else:
                self._rate_hint.configure(
                    text="= 金額(NT) ÷ 數量",
                    text_color="gray60")
        except (ValueError, ZeroDivisionError):
            pass

    def _fetch_price(self):
        """查詢當日市場行情（當日均價），不覆蓋交易匯率"""
        date = self.date_e.get().strip()
        cur  = self._get_currency()
        if not date or not cur:
            messagebox.showwarning("缺少資訊", "請先填入日期與幣種", parent=self)
            return
        self._fetch_btn.configure(state="disabled", text="查詢中…")

        def do_fetch():
            from api.price_fetcher import fetch_exchange_rate
            result = fetch_exchange_rate(cur, date)  # 不傳 qty/amt，僅查市場行情
            self.after(0, self._fill_market_price, result)

        threading.Thread(target=do_fetch, daemon=True).start()

    def _fill_market_price(self, result: dict):
        """只更新市場行情欄位（當日均價/高/低），不覆蓋交易匯率"""
        self._fetch_btn.configure(state="normal", text="查詢市場行情")
        if result.get("error"):
            messagebox.showerror("查詢失敗", result["error"], parent=self)
            return

        def _set_entry(e, val):
            if val is not None:
                e.delete(0, "end")
                e.insert(0, f"{val:,.4f}")

        # 只填當日均價，不動交易匯率
        _set_entry(self.avg_e, result.get("daily_avg"))
        h = result.get("daily_high")
        l = result.get("daily_low")
        if h and l:
            self.hl_lbl.configure(
                text=f"高 {h:,.2f} ／ 低 {l:,.2f}",
                text_color="#aaffaa")

    def _save(self):
        date = self.date_e.get().strip()
        time = self.time_e.get().strip()
        cur  = self._get_currency()
        if not date:
            messagebox.showwarning("缺少資料", "日期為必填欄位", parent=self)
            return
        if not cur:
            messagebox.showwarning("缺少資料", "幣種為必填欄位", parent=self)
            return

        def _float(e: ctk.CTkEntry):
            s = e.get().strip().replace(",", "")
            try: return float(s) if s else None
            except ValueError: return None

        data = {
            "id":            self.row.get("id"),
            "tx_date":       date,
            "tx_time":       time or "00:00:00",
            "from_addr":     self.from_e.get().strip(),
            "to_addr":       self.to_e.get().strip(),
            "amount_ntd":    _float(self.amt_e),
            "quantity":      _float(self.qty_e),
            "currency":      cur,
            "exchange_rate": _float(self.rate_e),
            "daily_avg":     _float(self.avg_e),
            "daily_high":    None,
            "daily_low":     None,
            "notes":         self.notes_e.get().strip(),
            "source_doc":    self.row.get("source_doc", ""),
        }

        try:
            new_id = _db.upsert_victim_transaction(self.case_id, data)
            data["id"] = new_id
        except Exception as e:
            messagebox.showerror("儲存失敗", str(e), parent=self)
            return
        if self.on_save:
            self.on_save(data)
        self.destroy()


# ── 被害人陳述交易記錄主面板 ──────────────────────────────────────────────────

class VictimTxPanel(ctk.CTkFrame):
    def __init__(self, parent, case_id: int, **kw):
        super().__init__(parent, corner_radius=8, **kw)
        self.case_id = case_id
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._build()
        self._load()

    def set_case(self, case_id: int):
        self.case_id = case_id
        self._load()

    def _build(self):
        # 工具列
        bar = ctk.CTkFrame(self, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=6, pady=(6, 2))

        ctk.CTkLabel(bar, text="被害人陳述交易紀錄",
                     font=("Microsoft JhengHei", 12, "bold")).pack(side="left")

        ctk.CTkButton(bar, text="＋ 新增", width=80,
                      font=("Microsoft JhengHei", 10),
                      fg_color="#1d6b3e",
                      command=self._add).pack(side="left", padx=(12, 2))
        ctk.CTkButton(bar, text="✎ 編輯", width=70,
                      font=("Microsoft JhengHei", 10),
                      fg_color="#2a4a8a",
                      command=self._edit).pack(side="left", padx=2)
        ctk.CTkButton(bar, text="刪除", width=60,
                      font=("Microsoft JhengHei", 10),
                      fg_color="#7a1f1f",
                      command=self._delete).pack(side="left", padx=2)
        ctk.CTkButton(bar, text="📂 從文件匯入", width=110,
                      font=("Microsoft JhengHei", 10),
                      fg_color="#4a3a7a",
                      command=self._import_doc).pack(side="left", padx=8)
        ctk.CTkButton(bar, text="💱 批次查匯率", width=110,
                      font=("Microsoft JhengHei", 10),
                      fg_color="#3a5a2a",
                      command=self._batch_fetch_price).pack(side="left", padx=2)
        ctk.CTkButton(bar, text="📊 匯出 Excel", width=100,
                      font=("Microsoft JhengHei", 10),
                      fg_color="#2d5a4f",
                      command=self._export_excel).pack(side="left", padx=2)

        self._count_lbl = ctk.CTkLabel(bar, text="",
                                        font=("Microsoft JhengHei", 10),
                                        text_color="gray60")
        self._count_lbl.pack(side="right", padx=8)

        # 表格
        tree_frame = ctk.CTkFrame(self, corner_radius=6)
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0, 6))
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        style = ttk.Style()
        style.configure("Treeview", background="#2b2b2b", foreground="white",
                        fieldbackground="#2b2b2b", rowheight=24,
                        font=("Consolas", 10))
        style.configure("Treeview.Heading", background="#1f1f2e",
                        foreground="white", font=("Microsoft JhengHei", 10, "bold"))
        style.map("Treeview", background=[("selected", "#1f538d")])

        col_ids = [c[0] for c in _COLS]
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        self._tree = ttk.Treeview(tree_frame, columns=col_ids,
                                   show="headings",
                                   yscrollcommand=vsb.set,
                                   xscrollcommand=hsb.set)
        vsb.config(command=self._tree.yview)
        hsb.config(command=self._tree.xview)

        for name, _, width in _COLS:
            self._tree.heading(name, text=name)
            self._tree.column(name, width=width, minwidth=50, stretch=False)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self._tree.bind("<Double-1>", lambda _: self._edit())

        self._rows: list[dict] = []
        self._add_context_menu()

    def _add_context_menu(self):
        menu = tk.Menu(self._tree, tearoff=0,
                       bg="#2b2b2b", fg="white",
                       activebackground="#1f538d",
                       activeforeground="white",
                       font=("Microsoft JhengHei", 10))
        menu.add_command(label="複製此格", command=self._copy_cell)
        menu.add_command(label="複製整列", command=self._copy_row)
        menu.add_separator()
        menu.add_command(label="編輯",     command=self._edit)
        menu.add_command(label="刪除",     command=self._delete)
        self._ctx_col = ""

        def show(e):
            iid = self._tree.identify_row(e.y)
            if iid:
                self._tree.selection_set(iid)
            col = self._tree.identify_column(e.x)
            try:
                self._ctx_col = _COLS[int(col.replace("#",""))-1][0]
            except Exception:
                self._ctx_col = ""
            menu.tk_popup(e.x_root, e.y_root)
        self._tree.bind("<Button-3>", show)

    def _copy_cell(self):
        sel = self._tree.selection()
        if not sel: return
        vals = self._tree.item(sel[0])["values"]
        cols = [c[0] for c in _COLS]
        try:
            idx = cols.index(self._ctx_col)
            self.clipboard_clear(); self.clipboard_append(str(vals[idx]))
        except Exception: pass

    def _copy_row(self):
        sel = self._tree.selection()
        if not sel: return
        vals = self._tree.item(sel[0])["values"]
        self.clipboard_clear()
        self.clipboard_append("\t".join(str(v) for v in vals))

    # ── 資料操作 ───────────────────────────────────────────────────────────────

    def _load(self):
        for iid in self._tree.get_children():
            self._tree.delete(iid)
        self._rows = _db.get_victim_transactions(self.case_id)
        for r in self._rows:
            self._tree.insert("", "end", iid=str(r["id"]), values=[
                r.get("tx_date", ""),
                r.get("tx_time", ""),
                r.get("from_addr", ""),
                r.get("to_addr",   ""),
                f"{r['amount_ntd']:,.2f}" if r.get("amount_ntd") else "",
                f"{r['quantity']:,.6f}"   if r.get("quantity")   else "",
                r.get("currency", ""),
                f"{r['exchange_rate']:,.2f}" if r.get("exchange_rate") else "",
                f"{r['daily_avg']:,.2f}"     if r.get("daily_avg")     else "",
            ])
        self._count_lbl.configure(text=f"共 {len(self._rows)} 筆")

    def _get_selected_row(self) -> dict | None:
        sel = self._tree.selection()
        if not sel:
            return None
        row_id = int(sel[0])
        return next((r for r in self._rows if r["id"] == row_id), None)

    def _add(self):
        VictimTxDialog(self, self.case_id, on_save=lambda _: self._load())

    def _edit(self):
        row = self._get_selected_row()
        if not row:
            messagebox.showinfo("提示", "請先選取一筆記錄")
            return
        VictimTxDialog(self, self.case_id, row=row, on_save=lambda _: self._load())

    def _delete(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("提示", "請先選取要刪除的記錄")
            return
        if not messagebox.askyesno("確認刪除",
                                   f"確定刪除選取的 {len(sel)} 筆記錄？"):
            return
        for iid in sel:
            _db.delete_victim_transaction(int(iid))
        self._load()

    def _import_doc(self):
        """從文件檔案提取交易記錄"""
        paths = filedialog.askopenfilenames(
            title="選擇文件檔案",
            filetypes=[
                ("支援文件", "*.pdf *.docx *.xlsx *.odt *.txt"),
                ("PDF", "*.pdf"), ("Word", "*.docx"),
                ("Excel", "*.xlsx"), ("ODT", "*.odt"),
                ("文字", "*.txt"), ("全部", "*.*"),
            ]
        )
        if not paths:
            return

        from analyzer.doc_transaction_extractor import (
            extract_text_from_file, _parse_transactions)

        imported = 0
        for path in paths:
            text = extract_text_from_file(path)
            if not text:
                continue
            txs = _parse_transactions(text)
            for t in txs:
                t["source_doc"] = path
                _db.upsert_victim_transaction(self.case_id, t)
                imported += 1

        self._load()
        messagebox.showinfo("匯入完成",
                            f"從 {len(paths)} 個文件共匯入 {imported} 筆交易記錄。\n"
                            "請逐一確認並修正資料。")

    def _batch_fetch_price(self):
        """批次查詢所有缺少匯率資料的記錄"""
        missing = [r for r in self._rows
                   if not r.get("daily_avg") and r.get("tx_date") and r.get("currency")]
        if not missing:
            messagebox.showinfo("已完整", "所有記錄都已有匯率資料")
            return
        if not messagebox.askyesno("批次查詢",
                                   f"將查詢 {len(missing)} 筆缺少匯率的記錄。\n"
                                   "（需連線至 CoinGecko，可能需要數秒）\n\n"
                                   "是否繼續？"):
            return

        def do_batch():
            from api.price_fetcher import calc_exchange_rate
            for r in missing:
                # 只查市場行情（當日均價），不覆蓋交易匯率
                result = fetch_exchange_rate(r["currency"], r["tx_date"])
                if result.get("daily_avg"):
                    r["daily_avg"]  = result["daily_avg"]
                    r["daily_high"] = result.get("daily_high")
                    r["daily_low"]  = result.get("daily_low")
                # 若交易匯率空白且有金額+數量，則由被害人陳述計算
                if not r.get("exchange_rate"):
                    r["exchange_rate"] = calc_exchange_rate(
                        r.get("amount_ntd"), r.get("quantity"))
                _db.upsert_victim_transaction(self.case_id, r)
            self.after(0, self._load)
            self.after(0, lambda: messagebox.showinfo(
                "查詢完成", f"已完成 {len(missing)} 筆匯率查詢"))

        threading.Thread(target=do_batch, daemon=True).start()

    def _export_excel(self):
        if not self._rows:
            messagebox.showinfo("無資料", "目前沒有交易記錄可匯出")
            return
        path = filedialog.asksaveasfilename(
            title="儲存 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"被害人交易記錄_case{self.case_id}.xlsx"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "被害人陳述交易紀錄"

            # 標題行
            headers = [c[0] for c in _COLS] + ["備註", "來源文件"]
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color="FFFFFF",
                                 name="Microsoft JhengHei")
                cell.fill = PatternFill("solid", fgColor="1F3864")
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[
                    chr(64 + ci)].width = max(len(h) + 4, 14)

            # 資料行
            for ri, r in enumerate(self._rows, 2):
                vals = [
                    r.get("tx_date",""),   r.get("tx_time",""),
                    r.get("from_addr",""), r.get("to_addr",""),
                    r.get("amount_ntd"),   r.get("quantity"),
                    r.get("currency",""),  r.get("exchange_rate"),
                    r.get("daily_avg"),    r.get("notes",""),
                    r.get("source_doc",""),
                ]
                for ci, v in enumerate(vals, 1):
                    ws.cell(row=ri, column=ci, value=v)

            wb.save(path)
            messagebox.showinfo("匯出完成", f"已儲存至：\n{path}")
        except Exception as e:
            messagebox.showerror("匯出失敗", str(e))
