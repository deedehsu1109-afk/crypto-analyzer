from __future__ import annotations
import csv
import os
import datetime
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _now_str() -> str:
    return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


# ── 共用：將 profile 摘要轉為列表 ──────────────────────────────────────────────

def _summary_rows(profile: dict) -> list[tuple[str, Any]]:
    chain = profile.get("chain", "")
    unit  = {"ETH": "ETH", "TRX": "TRX", "BTC": "BTC"}.get(chain, "")
    amt_key = {"ETH": "out_total_eth", "TRX": "out_total_trx", "BTC": "out_total_btc"}.get(chain, "")
    in_key  = {"ETH": "in_total_eth",  "TRX": "in_total_trx",  "BTC": "in_total_btc"}.get(chain, "")
    fee_key = {"ETH": "total_fee_eth", "TRX": "total_fee_trx", "BTC": "total_fee_btc"}.get(chain, "")

    rows = [
        ("區塊鏈", chain),
        ("錢包地址", profile.get("address", "")),
        ("首次交易時間", profile.get("first_tx_time", "N/A")),
        ("最後交易時間", profile.get("last_tx_time", "N/A")),
        ("首次資金來源", profile.get("first_source", "N/A")),
        ("發起交易次數", profile.get("out_count", 0)),
        (f"發起交易總金額 ({unit})", profile.get(amt_key, 0)),
        ("接受交易次數", profile.get("in_count", 0)),
        (f"接受交易總金額 ({unit})", profile.get(in_key, 0)),
        (f"總手續費 ({unit})", profile.get(fee_key, 0)),
        ("最多手續費流向地址", profile.get("top_fee_dest", "N/A")),
    ]
    if chain == "ETH":
        rows.append(("ERC-20 轉帳次數", profile.get("erc20_transfer_count", 0)))
    elif chain == "TRX":
        rows.append(("TRC-20 轉帳次數", profile.get("trc20_transfer_count", 0)))
    return rows


# ── Excel 匯出 ─────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Microsoft JhengHei", size=10)
_LABEL_FILL   = PatternFill("solid", fgColor="D6E4F0")
_LABEL_FONT   = Font(bold=True, name="Microsoft JhengHei", size=10)
_BODY_FONT    = Font(name="Microsoft JhengHei", size=10)
_THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _style_header(ws, row: int, col: int, value: str):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _THIN_BORDER


def _style_label(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _LABEL_FILL
    cell.font = _LABEL_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def _style_value(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _BODY_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def export_excel(profile: dict, output_dir: str = ".") -> str:
    wb = Workbook()

    # ── 分頁1：摘要 ──
    ws1 = wb.active
    ws1.title = "錢包摘要"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 55

    _style_header(ws1, 1, 1, "項目")
    _style_header(ws1, 1, 2, "內容")

    summary = _summary_rows(profile)
    for i, (label, value) in enumerate(summary, start=2):
        _style_label(ws1, i, 1, label)
        _style_value(ws1, i, 2, value)

    # ── 分頁2：授權對象 ──
    ws2 = wb.create_sheet("授權對象")
    approvals = profile.get("approval_targets", [])
    chain = profile.get("chain", "")
    if chain == "ETH":
        headers = ["合約地址", "授權對象 (Spender)", "交易 Hash", "授權時間"]
        ws2.column_dimensions["A"].width = 44
        ws2.column_dimensions["B"].width = 44
        ws2.column_dimensions["C"].width = 68
        ws2.column_dimensions["D"].width = 22
        for col, h in enumerate(headers, 1):
            _style_header(ws2, 1, col, h)
        for r, a in enumerate(approvals, 2):
            _style_value(ws2, r, 1, a.get("contract", ""))
            _style_value(ws2, r, 2, a.get("spender", ""))
            _style_value(ws2, r, 3, a.get("tx_hash", ""))
            _style_value(ws2, r, 4, a.get("time", ""))
    elif chain == "TRX":
        headers = ["合約地址", "授權對象 (Spender)", "授權金額"]
        ws2.column_dimensions["A"].width = 36
        ws2.column_dimensions["B"].width = 36
        ws2.column_dimensions["C"].width = 20
        for col, h in enumerate(headers, 1):
            _style_header(ws2, 1, col, h)
        for r, a in enumerate(approvals, 2):
            _style_value(ws2, r, 1, a.get("contract", ""))
            _style_value(ws2, r, 2, a.get("spender", ""))
            _style_value(ws2, r, 3, a.get("amount", ""))
    else:
        ws2.cell(row=1, column=1, value="此鏈不支援授權記錄")

    # ── 分頁3：原始交易 ──
    ws3 = wb.create_sheet("原始交易")
    raw_txs = profile.get("raw_txs", [])
    _write_raw_sheet(ws3, raw_txs, chain)

    # ── 分頁4：Token 轉帳 ──
    ws4 = wb.create_sheet("Token 轉帳")
    if chain == "ETH":
        _write_raw_sheet(ws4, profile.get("raw_erc20", []), "ERC20")
    elif chain == "TRX":
        _write_raw_sheet(ws4, profile.get("raw_trc20", []), "TRC20")
    else:
        ws4.cell(row=1, column=1, value="BTC 無 Token 轉帳資料")

    os.makedirs(output_dir, exist_ok=True)
    addr_short = profile.get("address", "unknown")[:10]
    chain_tag  = profile.get("chain", "XX")
    filename   = os.path.join(output_dir, f"{chain_tag}_{addr_short}_{_now_str()}.xlsx")
    wb.save(filename)
    return filename


def _write_raw_sheet(ws, txs: list[dict], chain: str):
    if not txs:
        ws.cell(row=1, column=1, value="無資料")
        return
    keys = list(txs[0].keys())
    for col, k in enumerate(keys, 1):
        _style_header(ws, 1, col, k)
        ws.column_dimensions[get_column_letter(col)].width = max(len(k) + 4, 14)
    for r, tx in enumerate(txs, 2):
        for col, k in enumerate(keys, 1):
            val = tx.get(k, "")
            if isinstance(val, (dict, list)):
                val = str(val)
            _style_value(ws, r, col, val)


# ── CSV 匯出 ───────────────────────────────────────────────────────────────────

def export_csv(profile: dict, output_dir: str = ".") -> list[str]:
    os.makedirs(output_dir, exist_ok=True)
    chain     = profile.get("chain", "XX")
    addr_short = profile.get("address", "unknown")[:10]
    files: list[str] = []

    # 摘要
    summary_file = os.path.join(output_dir, f"{chain}_{addr_short}_summary_{_now_str()}.csv")
    with open(summary_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["項目", "內容"])
        writer.writerows(_summary_rows(profile))
    files.append(summary_file)

    # 原始交易
    raw_txs = profile.get("raw_txs", [])
    if raw_txs:
        tx_file = os.path.join(output_dir, f"{chain}_{addr_short}_txs_{_now_str()}.csv")
        keys = list(raw_txs[0].keys())
        with open(tx_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(raw_txs)
        files.append(tx_file)

    return files
